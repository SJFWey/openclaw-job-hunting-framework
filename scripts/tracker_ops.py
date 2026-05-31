"""
tracker_ops.py - Job application tracker backed by an Excel workbook.

Sheets
------
Applications    : one row per application (APP-ID, company, position, status, ...)
Recommendations : one row per job lead recommendation (REC-ID, company, position, ...)
FollowUps       : one row per follow-up event (APP-ID, date, notes)

IDs
---
APP-YYYY-NNN  e.g. APP-2026-001
REC-YYYY-NNN  e.g. REC-2026-001
FU-YYYY-NNN   e.g. FU-2026-001
"""

from __future__ import annotations

import argparse
from collections import Counter
import datetime
import os
import sys
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError as exc:
    raise ImportError("openpyxl is required: pip install openpyxl") from exc

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

APP_STATUSES = [
    "draft",
    "submitted",
    "screening",
    "interview",
    "offer",
    "rejected",
    "withdrawn",
    "terminal",
]

REC_STATUSES = ["new", "accepted", "ignored", "rejected", "imported"]

APP_SHEET = "Applications"
REC_SHEET = "Recommendations"
FU_SHEET = "FollowUps"

APP_COLS = [
    "app_id",
    "company",
    "position",
    "source_url",
    "source_platform",
    "location",
    "status",
    "job_summary",
    "match_score",
    "fit_reasons",
    "created_at",
    "updated_at",
    "applied_at",
    "follow_up_due_at",
    "folder_path",
    "jd_path",
    "cover_letter_path",
    "packet_pdf_path",
    "notes",
]
REC_COLS = [
    "rec_id",
    "found_at",
    "company",
    "position",
    "url",
    "platform",
    "source_type",
    "location",
    "score",
    "technical_score",
    "location_score",
    "seniority_score",
    "risk_score",
    "seniority_fit",
    "domain_fit",
    "risk_flags",
    "reason",
    "decision_reason",
    "status",
    "linked_app_id",
]
FU_COLS = ["fu_id", "app_id", "fu_date", "channel", "notes"]

DEFAULT_PATH = Path("data/applications.xlsx")
MATERIALS_ROOT = Path(
    os.environ.get("JOB_HUNTER_MATERIALS_ROOT")
    or "materials/applications"
)

# ---------------------------------------------------------------------------
# Workbook bootstrap
# ---------------------------------------------------------------------------


def _today_str() -> str:
    return str(datetime.date.today())


def _row_to_dict(cols: list[str], row: tuple) -> dict:
    values = list(row)
    if len(values) < len(cols):
        values.extend([None] * (len(cols) - len(values)))
    return dict(zip(cols, values))


def _sheet_row_to_dict(ws, cols: list[str], row: tuple) -> dict:
    headers = _sheet_headers(ws)
    raw = _row_to_dict(headers, row)
    return {col: raw.get(col) for col in cols}


def _normalize_text(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def normalize_url(url: str) -> str:
    """Normalize a job URL for recommendation deduplication."""
    raw = str(url or "").strip()
    if not raw:
        return ""
    parsed = urlsplit(raw)
    if not parsed.scheme or not parsed.netloc:
        return raw.rstrip("/").casefold()
    query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.casefold().startswith("utm_")
        and key.casefold() not in {"fbclid", "gclid", "msclkid"}
    ]
    path = parsed.path.rstrip("/") or "/"
    return urlunsplit(
        (
            parsed.scheme.casefold(),
            parsed.netloc.casefold(),
            path,
            urlencode(query, doseq=True),
            "",
        )
    )


def _recommendation_dedupe_key(rec: dict) -> str:
    normalized_url = normalize_url(str(rec.get("url") or ""))
    if normalized_url:
        return f"url:{normalized_url}"
    company = _normalize_text(rec.get("company"))
    position = _normalize_text(rec.get("position"))
    location = _normalize_text(rec.get("location"))
    return f"job:{company}|{position}|{location}"


def _make_header_row(ws, cols: list[str]) -> None:
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")


def create_workbook(path: Path) -> None:
    """Create a fresh workbook with all sheets and headers."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws_app = wb.create_sheet(APP_SHEET)
    _make_header_row(ws_app, APP_COLS)

    ws_rec = wb.create_sheet(REC_SHEET)
    _make_header_row(ws_rec, REC_COLS)

    ws_fu = wb.create_sheet(FU_SHEET)
    _make_header_row(ws_fu, FU_COLS)

    wb.save(path)


def _load(path: Path) -> Workbook:
    path = Path(path)
    if not path.exists():
        create_workbook(path)
    wb = load_workbook(path)
    if _ensure_workbook_schema(wb):
        wb.save(path)
    return wb


def _ensure_workbook_schema(wb: Workbook) -> bool:
    """Normalize sheet headers for backward-compatible workbook upgrades."""
    changed = False
    expected_by_sheet = {
        APP_SHEET: APP_COLS,
        REC_SHEET: REC_COLS,
        FU_SHEET: FU_COLS,
    }
    for sheet, expected_cols in expected_by_sheet.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = _sheet_headers(ws)
        if headers == expected_cols:
            continue
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = _row_to_dict(headers, row)
            if any(value is not None for value in raw.values()):
                data_rows.append([raw.get(col) for col in expected_cols])
        ws.delete_rows(1, ws.max_row)
        _make_header_row(ws, expected_cols)
        for data_row in data_rows:
            ws.append(data_row)
        changed = True
    return changed


# ---------------------------------------------------------------------------
# ID generation helpers
# ---------------------------------------------------------------------------


def _next_id(ws, prefix: str, col_idx: int = 1) -> str:
    """Generate next sequential ID for a sheet."""
    year = datetime.date.today().year
    max_n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = row[col_idx - 1]
        if cell and str(cell).startswith(f"{prefix}-{year}-"):
            try:
                n = int(str(cell).split("-")[-1])
                max_n = max(max_n, n)
            except ValueError:
                pass
    return f"{prefix}-{year}-{max_n + 1:03d}"


def _find_row(ws, record_id: str, id_col: int = 1):
    for row in ws.iter_rows(min_row=2):
        if row[id_col - 1].value == record_id:
            return row
    return None


# ---------------------------------------------------------------------------
# Applications CRUD
# ---------------------------------------------------------------------------


def add_application(
    path: Path,
    company: str,
    position: str,
    source_url: str = "",
    source_platform: str = "",
    location: str = "",
    applied_date: str = "",
    follow_up_due_at: str = "",
    jd_path: str = "",
    job_summary: str = "",
    match_score: float = 0.0,
    fit_reasons: str = "",
    folder_path: str = "",
    cover_letter_path: str = "",
    packet_pdf_path: str = "",
    notes: str = "",
) -> str:
    """Add a new application row. Returns the new APP-ID."""
    wb = _load(path)
    ws = wb[APP_SHEET]
    app_id = _next_id(ws, "APP")
    today = _today_str()
    ws.append(
        [
            app_id,
            company,
            position,
            source_url,
            source_platform,
            location,
            "draft",
            job_summary,
            match_score,
            fit_reasons,
            today,
            today,
            applied_date,
            follow_up_due_at,
            folder_path,
            jd_path,
            cover_letter_path,
            packet_pdf_path,
            notes,
        ]
    )
    wb.save(path)
    return app_id


def update_application_status(path: Path, app_id: str, status: str) -> bool:
    """Update status of an existing application. Returns True if found."""
    if status not in APP_STATUSES:
        raise ValueError(f"Invalid status '{status}'. Choose from: {APP_STATUSES}")
    wb = _load(path)
    ws = wb[APP_SHEET]
    status_idx = APP_COLS.index("status")
    updated_at_idx = APP_COLS.index("updated_at")
    row = _find_row(ws, app_id)
    if row is None:
        return False
    row[status_idx].value = status
    row[updated_at_idx].value = _today_str()
    wb.save(path)
    return True


def update_application_paths(
    path: Path,
    app_id: str,
    folder_path: Optional[str] = None,
    jd_path: Optional[str] = None,
    cover_letter_path: Optional[str] = None,
    packet_pdf_path: Optional[str] = None,
) -> bool:
    """Update path fields on an existing application. Returns True if found."""
    wb = _load(path)
    ws = wb[APP_SHEET]
    row = _find_row(ws, app_id)
    if row is None:
        return False

    updates = {
        "folder_path": folder_path,
        "jd_path": jd_path,
        "cover_letter_path": cover_letter_path,
        "packet_pdf_path": packet_pdf_path,
    }
    changed = False
    for field, value in updates.items():
        if value is not None:
            row[APP_COLS.index(field)].value = value
            changed = True
    if changed:
        row[APP_COLS.index("updated_at")].value = _today_str()
    wb.save(path)
    return True


def get_application(path: Path, app_id: str) -> Optional[dict]:
    """Return application dict or None."""
    wb = _load(path)
    ws = wb[APP_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == app_id:
            return _row_to_dict(APP_COLS, row)
    return None


def list_applications(path: Path, status: Optional[str] = None) -> list[dict]:
    """Return list of application dicts, optionally filtered by status."""
    wb = _load(path)
    ws = wb[APP_SHEET]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = _row_to_dict(APP_COLS, row)
        if status is None or d.get("status") == status:
            rows.append(d)
    return rows


def mark_terminal(path: Path, app_id: str) -> bool:
    """Mark an application as terminal (rejected/withdrawn/offer accepted)."""
    return update_application_status(path, app_id, "terminal")


# ---------------------------------------------------------------------------
# Recommendations CRUD
# ---------------------------------------------------------------------------


def add_recommendation(
    path: Path,
    company: str,
    position: str,
    url: str = "",
    platform: str = "",
    source_type: str = "",
    location: str = "",
    score: float = 0.0,
    technical_score: float = 0.0,
    location_score: float = 0.0,
    seniority_score: float = 0.0,
    risk_score: float = 0.0,
    seniority_fit: str = "",
    domain_fit: str = "",
    risk_flags: str = "",
    reason: str = "",
    decision_reason: str = "",
) -> str:
    """Add a job lead recommendation. Returns REC-ID."""
    wb = _load(path)
    ws_rec = wb[REC_SHEET]
    rec_id = _next_id(ws_rec, "REC")
    found_at = _today_str()
    ws_rec.append(
        [
            rec_id,
            found_at,
            company,
            position,
            url,
            platform,
            source_type,
            location,
            score,
            technical_score,
            location_score,
            seniority_score,
            risk_score,
            seniority_fit,
            domain_fit,
            risk_flags,
            reason,
            decision_reason,
            "new",
            "",
        ]
    )
    wb.save(path)
    return rec_id


def update_recommendation_status(
    path: Path,
    rec_id: str,
    status: str,
    decision_reason: str = "",
) -> bool:
    if status not in REC_STATUSES:
        raise ValueError(f"Invalid status '{status}'. Choose from: {REC_STATUSES}")
    wb = _load(path)
    ws = wb[REC_SHEET]
    status_idx = REC_COLS.index("status")
    decision_reason_idx = REC_COLS.index("decision_reason")
    row = _find_row(ws, rec_id)
    if row is None:
        return False
    row[status_idx].value = status
    if decision_reason:
        row[decision_reason_idx].value = decision_reason
    wb.save(path)
    return True


def clear_recommendations(path: Path) -> int:
    """Remove all recommendation data rows while preserving workbook structure."""
    wb = _load(path)
    ws = wb[REC_SHEET]
    count = max(ws.max_row - 1, 0)
    if count:
        ws.delete_rows(2, count)
    wb.save(path)
    return count


def find_duplicate_recommendations(path: Path) -> list[dict]:
    """Return duplicate recommendation groups using URL, then company/title/location."""
    recs = list_recommendations(path)
    groups: dict[str, list[dict]] = {}
    for rec in recs:
        if _is_resolved_duplicate_recommendation(rec):
            continue
        key = _recommendation_dedupe_key(rec)
        if key.endswith("||"):
            continue
        groups.setdefault(key, []).append(rec)
    duplicates = []
    for key, items in groups.items():
        if len(items) > 1:
            duplicates.append(
                {
                    "key": key,
                    "keep": items[0].get("rec_id"),
                    "duplicates": [item.get("rec_id") for item in items[1:]],
                }
            )
    return duplicates


def dedupe_recommendations(path: Path, dry_run: bool = True) -> list[dict]:
    """
    Deduplicate recommendations. Keeps the first row for each duplicate key.

    Returns duplicate groups whether or not rows are deleted.
    """
    duplicates = find_duplicate_recommendations(path)
    if dry_run or not duplicates:
        return duplicates

    duplicate_ids = {
        rec_id
        for group in duplicates
        for rec_id in group["duplicates"]
        if rec_id
    }
    wb = _load(path)
    ws = wb[REC_SHEET]
    for row_idx in range(ws.max_row, 1, -1):
        rec_id = ws.cell(row_idx, 1).value
        if rec_id in duplicate_ids:
            ws.delete_rows(row_idx, 1)
    wb.save(path)
    return duplicates


def _is_resolved_duplicate_recommendation(rec: dict) -> bool:
    """Return True for duplicate rows intentionally retained for audit history."""
    status = str(rec.get("status") or "").strip().lower()
    reason = str(rec.get("decision_reason") or "").strip().lower()
    return status == "rejected" and reason.startswith("duplicate of rec-")


def promote_recommendation(path: Path, rec_id: str) -> bool:
    """Promote a recommendation to 'accepted' status."""
    return update_recommendation_status(path, rec_id, "accepted")


def promote_recommendation_to_application(path: Path, rec_id: str) -> Optional[str]:
    """
    Promote a job lead recommendation to a full application entry.

    - Looks up the recommendation by rec_id.
    - If rec already has a linked_app_id, return that existing app_id (idempotent).
    - Otherwise, calls add_application using rec's company/position/url/platform/location.
    - Updates rec status to 'imported' and sets linked_app_id.
    - Returns the new (or existing) app_id, or None if rec_id not found.
    """
    wb = _load(path)
    ws = wb[REC_SHEET]
    recommendation = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == rec_id:
            recommendation = _row_to_dict(REC_COLS, row)
            break

    if recommendation is None:
        return None

    linked_app_id = recommendation.get("linked_app_id")
    if linked_app_id:
        return str(linked_app_id)

    app_id = add_application(
        path,
        company=str(recommendation.get("company") or ""),
        position=str(recommendation.get("position") or ""),
        source_url=str(recommendation.get("url") or ""),
        source_platform=str(recommendation.get("platform") or ""),
        location=str(recommendation.get("location") or ""),
        match_score=float(recommendation.get("score") or 0.0),
        fit_reasons=str(recommendation.get("reason") or ""),
    )

    wb = _load(path)
    ws = wb[REC_SHEET]
    row = _find_row(ws, rec_id)
    if row is not None:
        row[REC_COLS.index("status")].value = "imported"
        row[REC_COLS.index("linked_app_id")].value = app_id
        wb.save(path)
    return app_id


def list_recommendations(path: Path, app_id: Optional[str] = None) -> list[dict]:
    wb = _load(path)
    ws = wb[REC_SHEET]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = _row_to_dict(REC_COLS, row)
        if app_id is None or d.get("linked_app_id") == app_id:
            rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# Workbook validation and reporting
# ---------------------------------------------------------------------------


def _sheet_headers(ws) -> list[str]:
    return [cell.value for cell in ws[1]]


def _validate_sheet_schema(wb, sheet: str, expected_cols: list[str]) -> list[str]:
    issues = []
    if sheet not in wb.sheetnames:
        return [f"missing sheet: {sheet}"]
    headers = _sheet_headers(wb[sheet])
    if headers[: len(expected_cols)] != expected_cols:
        issues.append(f"{sheet}: header mismatch")
        missing = [col for col in expected_cols if col not in headers]
        if missing:
            issues.append(f"{sheet}: missing headers: {', '.join(missing)}")
    return issues


def _path_exists_or_empty(value: object) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return True
    return Path(raw).exists()


def _materials_root_available(root: Path = MATERIALS_ROOT) -> bool:
    root = Path(root)
    return root.exists()


def _path_uses_materials_root(value: object, root: Path = MATERIALS_ROOT) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    try:
        return Path(raw).resolve().is_relative_to(Path(root).resolve())
    except OSError:
        return str(Path(raw)).startswith(str(root))


def validate_workbook(path: Path) -> list[str]:
    """Return validation issues for the tracker workbook."""
    issues: list[str] = []
    wb = _load(path)
    issues.extend(_validate_sheet_schema(wb, APP_SHEET, APP_COLS))
    issues.extend(_validate_sheet_schema(wb, REC_SHEET, REC_COLS))
    issues.extend(_validate_sheet_schema(wb, FU_SHEET, FU_COLS))
    if issues:
        return issues

    apps = list_applications(path)
    recs = list_recommendations(path)
    app_ids = {app.get("app_id") for app in apps}

    for app in apps:
        app_id = app.get("app_id")
        status = app.get("status")
        if status not in APP_STATUSES:
            issues.append(f"{app_id}: invalid application status: {status}")
        for field in ("folder_path", "jd_path", "cover_letter_path", "packet_pdf_path"):
            if not _path_exists_or_empty(app.get(field)):
                issues.append(f"{app_id}: missing path for {field}: {app.get(field)}")

    for rec in recs:
        rec_id = rec.get("rec_id")
        status = rec.get("status")
        if status not in REC_STATUSES:
            issues.append(f"{rec_id}: invalid recommendation status: {status}")
        if not str(rec.get("url") or "").strip():
            issues.append(f"{rec_id}: empty recommendation url")
        linked_app_id = rec.get("linked_app_id")
        if linked_app_id and linked_app_id not in app_ids:
            issues.append(f"{rec_id}: linked app does not exist: {linked_app_id}")

    for group in find_duplicate_recommendations(path):
        dupes = ", ".join(str(item) for item in group["duplicates"])
        issues.append(f"duplicate recommendations for {group['key']}: keep {group['keep']}, duplicates {dupes}")

    stale_new = [
        rec.get("rec_id")
        for rec in recs
        if rec.get("status") == "new"
        and rec.get("found_at")
        and _days_since(rec.get("found_at")) is not None
        and _days_since(rec.get("found_at")) >= 30
    ]
    if stale_new:
        issues.append(f"stale new recommendations >=30 days: {', '.join(str(i) for i in stale_new)}")

    uses_materials_root = any(
        _path_uses_materials_root(app.get(field))
        for app in apps
        for field in ("folder_path", "jd_path", "cover_letter_path", "packet_pdf_path")
    )
    if uses_materials_root and not _materials_root_available():
        issues.append(f"Materials root unavailable: {MATERIALS_ROOT}")

    return issues


def _days_since(raw_date: object) -> Optional[int]:
    try:
        value = datetime.date.fromisoformat(str(raw_date))
    except (TypeError, ValueError):
        return None
    return (datetime.date.today() - value).days


def tracker_stats(path: Path) -> dict:
    """Return basic tracker counts for applications, recommendations, and follow-ups."""
    apps = list_applications(path)
    recs = list_recommendations(path)
    fus = list_followups(path)
    return {
        "applications": len(apps),
        "applications_by_status": dict(Counter(str(app.get("status") or "") for app in apps)),
        "recommendations": len(recs),
        "recommendations_by_status": dict(Counter(str(rec.get("status") or "") for rec in recs)),
        "recommendations_by_source_type": dict(Counter(str(rec.get("source_type") or "") for rec in recs)),
        "recommendations_by_domain_fit": dict(Counter(str(rec.get("domain_fit") or "") for rec in recs)),
        "recommendations_by_seniority_fit": dict(Counter(str(rec.get("seniority_fit") or "") for rec in recs)),
        "followups": len(fus),
        "due_followups_7d": len(due_followups(path, days_since=7)),
    }


# ---------------------------------------------------------------------------
# Follow-up CRUD
# ---------------------------------------------------------------------------


def add_followup(path: Path, app_id: str, channel: str = "email", notes: str = "") -> str:
    """Log a follow-up event. Returns FU-ID."""
    wb = _load(path)
    ws = wb[FU_SHEET]
    fu_id = _next_id(ws, "FU")
    fu_date = str(datetime.date.today())
    ws.append([fu_id, app_id, fu_date, channel, notes])
    wb.save(path)
    return fu_id


def list_followups(path: Path, app_id: Optional[str] = None) -> list[dict]:
    wb = _load(path)
    ws = wb[FU_SHEET]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(FU_COLS, row))
        if app_id is None or d.get("app_id") == app_id:
            rows.append(d)
    return rows


def due_followups(path: Path, days_since: int = 7) -> list[dict]:
    """Return follow-up items for applications with no follow-up in N days."""
    cutoff = datetime.date.today() - datetime.timedelta(days=days_since)
    apps = list_applications(path, status=None)
    active = {
        a["app_id"]
        for a in apps
        if a.get("status") not in ("terminal", "rejected", "withdrawn")
    }
    fus = list_followups(path)
    last_fu: dict[str, datetime.date] = {}
    for fu in fus:
        aid = fu.get("app_id")
        raw = fu.get("fu_date")
        if aid and raw:
            try:
                d = datetime.date.fromisoformat(str(raw))
                if aid not in last_fu or d > last_fu[aid]:
                    last_fu[aid] = d
            except ValueError:
                pass
    due = []
    for aid in active:
        last = last_fu.get(aid)
        if last is None or last <= cutoff:
            app = get_application(path, aid)
            if app:
                due.append(app)
    return due


# ---------------------------------------------------------------------------
# JD markdown helper
# ---------------------------------------------------------------------------


def save_jd_markdown(app_id: str, content: str, base_dir: Optional[Path] = None) -> Path:
    """Save JD as markdown. Defaults to MATERIALS_ROOT/<app_id>/JD.md if base_dir is None."""
    if base_dir is None:
        base_dir = MATERIALS_ROOT / app_id
    else:
        base_dir = Path(base_dir)
    base_dir.mkdir(parents=True, exist_ok=True)
    out = base_dir / "JD.md"
    out.write_text(content, encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_cli() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="tracker_ops",
        description="Job application tracker CLI",
    )
    p.add_argument(
        "--db",
        default=str(DEFAULT_PATH),
        help="Path to the Excel workbook (default: data/applications.xlsx)",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    # init
    sub.add_parser("init", help="Create a fresh workbook")

    # add-app
    ap = sub.add_parser("add-app", help="Add a new application")
    ap.add_argument("company")
    ap.add_argument("position")
    ap.add_argument("--source-url", default="")
    ap.add_argument("--source-platform", "--platform", dest="source_platform", default="")
    ap.add_argument("--location", default="")
    ap.add_argument("--applied-date", default="")
    ap.add_argument("--follow-up-due-at", default="")
    ap.add_argument("--jd-path", default="")
    ap.add_argument("--job-summary", default="")
    ap.add_argument("--match-score", type=float, default=0.0)
    ap.add_argument("--fit-reasons", default="")
    ap.add_argument("--folder-path", default="")
    ap.add_argument("--cover-letter-path", default="")
    ap.add_argument("--packet-pdf-path", default="")
    ap.add_argument("--notes", default="")

    # update-app
    up = sub.add_parser("update-app", help="Update application status")
    up.add_argument("app_id")
    up.add_argument("status", choices=APP_STATUSES)

    # update-paths
    paths = sub.add_parser("update-paths", help="Update application file path fields")
    paths.add_argument("app_id")
    paths.add_argument("--folder-path", default=None)
    paths.add_argument("--jd-path", default=None)
    paths.add_argument("--cover-letter-path", default=None)
    paths.add_argument("--packet-pdf-path", default=None)

    # list-apps
    la = sub.add_parser("list-apps", help="List applications")
    la.add_argument("--status", default=None)

    # add-rec
    ar = sub.add_parser("add-rec", help="Add a recommendation")
    ar.add_argument("company")
    ar.add_argument("position")
    ar.add_argument("--url", default="")
    ar.add_argument("--platform", default="")
    ar.add_argument("--source-type", default="")
    ar.add_argument("--location", default="")
    ar.add_argument("--score", type=float, default=0.0)
    ar.add_argument("--technical-score", type=float, default=0.0)
    ar.add_argument("--location-score", type=float, default=0.0)
    ar.add_argument("--seniority-score", type=float, default=0.0)
    ar.add_argument("--risk-score", type=float, default=0.0)
    ar.add_argument("--seniority-fit", default="")
    ar.add_argument("--domain-fit", default="")
    ar.add_argument("--risk-flags", default="")
    ar.add_argument("--reason", default="")
    ar.add_argument("--decision-reason", default="")

    # list-recs
    lr = sub.add_parser("list-recs", help="List recommendations")
    lr.add_argument("--app-id", default=None)

    # update-rec-status
    urs = sub.add_parser("update-rec-status", help="Update recommendation status")
    urs.add_argument("rec_id")
    urs.add_argument("status", choices=REC_STATUSES)
    urs.add_argument("--decision-reason", default="")

    # clear-recs
    sub.add_parser("clear-recs", help="Delete all recommendation data rows")

    # dedupe-recs
    dr = sub.add_parser("dedupe-recs", help="Find or remove duplicate recommendations")
    dr.add_argument("--dry-run", action="store_true", help="Report duplicates without deleting")
    dr.add_argument("--apply", action="store_true", help="Delete duplicate rows after the first match")

    # validate
    sub.add_parser("validate", help="Validate workbook schema, statuses, paths, and duplicates")

    # stats
    sub.add_parser("stats", help="Print tracker statistics")

    # promote-rec
    pr = sub.add_parser("promote-rec", help="Promote a recommendation")
    pr.add_argument("rec_id")

    # promote-rec-to-app
    pra = sub.add_parser(
        "promote-rec-to-app",
        help="Promote a recommendation into an application",
    )
    pra.add_argument("rec_id")

    # add-followup
    af = sub.add_parser("add-followup", help="Log a follow-up")
    af.add_argument("app_id")
    af.add_argument("--channel", default="email")
    af.add_argument("--notes", default="")

    # due-followups
    df = sub.add_parser("due-followups", help="List applications due for follow-up")
    df.add_argument("--days", type=int, default=7)

    return p


def main(argv: list[str] | None = None) -> None:
    parser = _build_cli()
    args = parser.parse_args(argv)
    db = Path(args.db)

    if args.cmd == "init":
        create_workbook(db)
        print(f"Created: {db}")

    elif args.cmd == "add-app":
        app_id = add_application(
            db,
            args.company,
            args.position,
            source_url=args.source_url,
            source_platform=args.source_platform,
            location=args.location,
            applied_date=args.applied_date,
            follow_up_due_at=args.follow_up_due_at,
            jd_path=args.jd_path,
            job_summary=args.job_summary,
            match_score=args.match_score,
            fit_reasons=args.fit_reasons,
            folder_path=args.folder_path,
            cover_letter_path=args.cover_letter_path,
            packet_pdf_path=args.packet_pdf_path,
            notes=args.notes,
        )
        print(f"Added application: {app_id}")

    elif args.cmd == "update-app":
        found = update_application_status(db, args.app_id, args.status)
        if found:
            print(f"Updated {args.app_id} -> {args.status}")
        else:
            print(f"Not found: {args.app_id}", file=sys.stderr)
            sys.exit(1)

    elif args.cmd == "update-paths":
        found = update_application_paths(
            db,
            args.app_id,
            folder_path=args.folder_path,
            jd_path=args.jd_path,
            cover_letter_path=args.cover_letter_path,
            packet_pdf_path=args.packet_pdf_path,
        )
        if found:
            print(f"Updated paths for {args.app_id}")
        else:
            print(f"Not found: {args.app_id}", file=sys.stderr)
            sys.exit(1)

    elif args.cmd == "list-apps":
        apps = list_applications(db, status=args.status)
        for a in apps:
            print(
                f"{a['app_id']}  {a['company']:<25} {a['position']:<30} {a['status']}"
            )

    elif args.cmd == "add-rec":
        rec_id = add_recommendation(
            db,
            args.company,
            args.position,
            url=args.url,
            platform=args.platform,
            source_type=args.source_type,
            location=args.location,
            score=args.score,
            technical_score=args.technical_score,
            location_score=args.location_score,
            seniority_score=args.seniority_score,
            risk_score=args.risk_score,
            seniority_fit=args.seniority_fit,
            domain_fit=args.domain_fit,
            risk_flags=args.risk_flags,
            reason=args.reason,
            decision_reason=args.decision_reason,
        )
        print(f"Added recommendation: {rec_id}")

    elif args.cmd == "list-recs":
        recs = list_recommendations(db, app_id=args.app_id)
        for r in recs:
            linked = f" -> {r['linked_app_id']}" if r.get("linked_app_id") else ""
            score = r.get("score") if r.get("score") is not None else ""
            domain = r.get("domain_fit") or ""
            print(
                f"{r['rec_id']}  {r['company']:<25} {r['position']:<30} {r['status']:<8} {score!s:<4} {domain}{linked}"
            )

    elif args.cmd == "update-rec-status":
        found = update_recommendation_status(
            db,
            args.rec_id,
            args.status,
            decision_reason=args.decision_reason,
        )
        if found:
            print(f"Updated {args.rec_id} -> {args.status}")
        else:
            print(f"Not found: {args.rec_id}", file=sys.stderr)
            sys.exit(1)

    elif args.cmd == "clear-recs":
        count = clear_recommendations(db)
        print(f"Deleted recommendations: {count}")

    elif args.cmd == "dedupe-recs":
        if args.dry_run and args.apply:
            print("Choose either --dry-run or --apply, not both.", file=sys.stderr)
            sys.exit(2)
        dry_run = not args.apply
        duplicates = dedupe_recommendations(db, dry_run=dry_run)
        if not duplicates:
            print("No duplicate recommendations found.")
        else:
            action = "Would delete" if dry_run else "Deleted"
            for group in duplicates:
                dupes = ", ".join(str(item) for item in group["duplicates"])
                print(f"{action}: {dupes} (keep {group['keep']}; key {group['key']})")

    elif args.cmd == "validate":
        issues = validate_workbook(db)
        if issues:
            for issue in issues:
                print(f"ERROR: {issue}")
            sys.exit(1)
        print("Tracker validation OK.")

    elif args.cmd == "stats":
        stats = tracker_stats(db)
        print(f"Applications: {stats['applications']}")
        print(f"Applications by status: {stats['applications_by_status']}")
        print(f"Recommendations: {stats['recommendations']}")
        print(f"Recommendations by status: {stats['recommendations_by_status']}")
        print(f"Recommendations by source type: {stats['recommendations_by_source_type']}")
        print(f"Recommendations by domain fit: {stats['recommendations_by_domain_fit']}")
        print(f"Recommendations by seniority fit: {stats['recommendations_by_seniority_fit']}")
        print(f"Follow-ups: {stats['followups']}")
        print(f"Due follow-ups (7d): {stats['due_followups_7d']}")

    elif args.cmd == "promote-rec":
        found = promote_recommendation(db, args.rec_id)
        if found:
            print(f"Accepted recommendation: {args.rec_id}")
        else:
            print(f"Not found: {args.rec_id}", file=sys.stderr)
            sys.exit(1)

    elif args.cmd == "promote-rec-to-app":
        app_id = promote_recommendation_to_application(db, args.rec_id)
        if app_id is not None:
            print(f"Promoted {args.rec_id} -> {app_id}")
        else:
            print(f"Not found: {args.rec_id}", file=sys.stderr)
            sys.exit(1)

    elif args.cmd == "add-followup":
        fu_id = add_followup(db, args.app_id, args.channel, args.notes)
        print(f"Logged follow-up: {fu_id}")

    elif args.cmd == "due-followups":
        due = due_followups(db, days_since=args.days)
        for a in due:
            print(f"{a['app_id']}  {a['company']:<25} last follow-up overdue")


if __name__ == "__main__":
    main()
