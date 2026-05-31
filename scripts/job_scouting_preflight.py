#!/usr/bin/env python3
"""Pre-run context collector for the daily job-scouting cron job.

Runs before the LLM agent. Stdout is injected into the cron prompt as context.
The **last non-empty line** must be JSON wake-gate metadata, e.g.
``{"wakeAgent": true}`` or ``{"wakeAgent": false, "reason": "..."}``.
"""
from __future__ import annotations

import datetime as dt
import json
import os
import subprocess
import textwrap
import argparse
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

WORKDIR = Path(
    os.environ.get("JOB_HUNTER_WORKDIR")
    or Path(__file__).resolve().parent.parent
).resolve()
TRACKER = Path(os.environ.get("JOB_HUNTER_TRACKER", WORKDIR / "data" / "applications.xlsx"))
KEYWORD_PACK = Path(
    os.environ.get("JOB_HUNTER_KEYWORD_PACK", WORKDIR / "docs" / "agent_job_search_keywords.json")
)
TUNING_CONFIG = Path(
    os.environ.get("JOB_HUNTER_TUNING_CONFIG", WORKDIR / "docs" / "job_scout_tuning.yaml")
)
FEEDBACK_LOG = Path(
    os.environ.get("JOB_HUNTER_FEEDBACK_LOG", WORKDIR / "data" / "job_scout_feedback.jsonl")
)
TARGET_COMPANIES = Path(
    os.environ.get("JOB_HUNTER_TARGET_COMPANIES", WORKDIR / "docs" / "target_companies.md")
)
SEARCH_DIAGNOSTICS = Path(
    os.environ.get("JOB_HUNTER_SEARCH_DIAGNOSTICS", WORKDIR / "docs" / "search_diagnostics.md")
)
OPENCLAW_HOME = Path(os.environ.get("OPENCLAW_HOME", Path.home() / ".openclaw"))
OPENCLAW_CRON_JOB_ID = os.environ.get("OPENCLAW_JOB_SCOUT_CRON_ID", "95f0899bca1d")
OPENCLAW_CRON_JOB_NAME = os.environ.get("OPENCLAW_JOB_SCOUT_CRON_NAME", "Daily job scouting")
OPENCLAW_CRON_JOBS = Path(
    os.environ.get("OPENCLAW_CRON_JOBS", OPENCLAW_HOME / "cron" / "jobs.json")
)
OPENCLAW_CRON_RUNS = Path(
    os.environ.get(
        "OPENCLAW_CRON_RUNS",
        OPENCLAW_HOME / "cron" / "runs" / f"{OPENCLAW_CRON_JOB_ID}.jsonl",
    )
)
RECENT_REC_LIMIT = 25
RECENT_KEY_LIMIT = 18
STATE_PATH = WORKDIR / ".state" / "job_scouting_preflight.state.json"
DUPLICATE_WINDOW_MINUTES = 45

sys.path.insert(0, str(WORKDIR / "scripts"))
try:
    from job_scout_tuning import (  # type: ignore
        format_tuning_summary,
        load_recent_feedback,
        load_tuning,
        summarize_feedback,
    )
    from search_plan import build_search_plan, format_markdown as format_search_plan_markdown  # type: ignore
except Exception as exc:  # pragma: no cover - preflight must degrade gracefully
    load_tuning = None
    format_tuning_summary = None
    load_recent_feedback = None
    summarize_feedback = None
    build_search_plan = None
    format_search_plan_markdown = None
    TUNING_IMPORT_ERROR = exc
else:
    TUNING_IMPORT_ERROR = None


def run(cmd: list[str], timeout: int = 120) -> tuple[bool, str]:
    env = os.environ.copy()
    env.pop("VIRTUAL_ENV", None)
    proc = subprocess.run(
        cmd,
        cwd=WORKDIR,
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )
    out = proc.stdout.strip()
    if proc.returncode != 0:
        return False, f"COMMAND FAILED ({proc.returncode}): {' '.join(cmd)}\n{out}"
    return True, out


def _load_state() -> dict:
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def _save_state(payload: dict) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _load_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def _load_tuning() -> dict:
    if load_tuning is None:
        return {
            "preset": "balanced",
            "config_status": f"import failed: {TUNING_IMPORT_ERROR}",
            "budgets": {
                "total_queries": 10,
                "p0_queries": 5,
                "p1_queries": 3,
                "p2_queries": 2,
                "target_companies_per_run": 6,
                "validated_target_min": 3,
                "validated_target_max": 8,
                "manual_check_target_min": 15,
                "manual_check_target_max": 25,
            },
            "freshness": {
                "prefer_posted_within_days": 14,
                "allow_unknown_date": True,
                "recheck_seen_after_days": 21,
            },
            "breadth": {
                "adjacent_role_ratio": 0.35,
                "broad_junior_ratio": 0.15,
                "far_south_requires_score": 0.75,
            },
            "depth": {
                "validation_mode": "standard",
                "require_full_jd_for_save": True,
                "extract_structured_signals": True,
                "max_full_jd_validations": 10,
            },
            "scoring": {
                "save_threshold": 0.5,
                "high_fit_threshold": 0.8,
                "manual_review_on_flags": True,
            },
            "feedback": {
                "use_recent_feedback": True,
                "feedback_window_days": 30,
                "min_signals_for_memory_proposal": 3,
            },
            "sources": {},
        }
    return load_tuning(TUNING_CONFIG)


def _cycle(items: list[str], count: int, offset: int) -> list[str]:
    if not items:
        return []
    count = min(count, len(items))
    return [items[(offset + i) % len(items)] for i in range(count)]


def _target_company_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    try:
        text = TARGET_COMPANIES.read_text(encoding="utf-8")
    except OSError:
        return rows
    for line in text.splitlines():
        if not line.startswith("|") or "---" in line or "Company" in line:
            continue
        parts = [part.strip() for part in line.strip("|").split("|")]
        if len(parts) < 4 or not parts[0]:
            continue
        rows.append(
            {
                "company": parts[0],
                "fit": parts[1],
                "region": parts[2],
                "hints": parts[3],
            }
        )
    return rows


def _recent_recommendation_keys() -> str:
    if not TRACKER.exists():
        return "Tracker missing; no recent recommendation keys available."

    code = f"""
from openpyxl import load_workbook
from pathlib import Path
p = Path({str(TRACKER)!r})
wb = load_workbook(p, data_only=True)
ws = wb['Recommendations']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
idx = {{h: i + 1 for i, h in enumerate(headers)}}
rows = []
for r in range(2, ws.max_row + 1):
    rec_id = ws.cell(r, idx['rec_id']).value
    if not rec_id:
        continue
    rows.append((
        rec_id,
        ws.cell(r, idx['company']).value,
        ws.cell(r, idx['position']).value,
        ws.cell(r, idx['location']).value,
        ws.cell(r, idx['status']).value,
    ))
for rec_id, company, position, location, status in rows[-{RECENT_KEY_LIMIT}:]:
    print(f"{{rec_id}} | {{status}} | {{company}} | {{position}} | {{location}}")
"""
    ok, out = run(["uv", "run", "python", "-c", code], timeout=60)
    return out if ok else out


def _recent_diagnostics() -> str:
    try:
        lines = SEARCH_DIAGNOSTICS.read_text(encoding="utf-8").splitlines()
    except OSError:
        return "No search diagnostics file found yet."
    useful = [line for line in lines if line.startswith("- ")]
    if not useful:
        return "Search diagnostics file exists, but no bullet notes were found."
    return "\n".join(useful[-12:])


def _feedback_summary(tuning: dict) -> str:
    feedback_cfg = tuning.get("feedback") if isinstance(tuning.get("feedback"), dict) else {}
    if not feedback_cfg.get("use_recent_feedback", True):
        return "Recent feedback disabled by tuning config."
    if load_recent_feedback is None or summarize_feedback is None:
        return f"Feedback unavailable: tuning import failed ({TUNING_IMPORT_ERROR})"
    rows = load_recent_feedback(
        FEEDBACK_LOG,
        window_days=int(feedback_cfg.get("feedback_window_days") or 30),
        today=dt.date.today(),
    )
    return summarize_feedback(
        rows,
        min_signals=int(feedback_cfg.get("min_signals_for_memory_proposal") or 3),
    )


def _daily_search_plan(now: dt.datetime, tuning: dict) -> str:
    if build_search_plan is None or format_search_plan_markdown is None:
        return f"Search plan unavailable: tuning/search import failed ({TUNING_IMPORT_ERROR})"
    plan = build_search_plan(
        date=now.date(),
        keyword_pack=_load_json(KEYWORD_PACK),
        tuning=tuning,
    )
    return format_search_plan_markdown(plan, include_rollout_notes=True)


def _latest_run_delivery_line() -> str:
    if not OPENCLAW_CRON_RUNS.exists():
        return f"No cron run history found: {OPENCLAW_CRON_RUNS}"
    last: dict | None = None
    try:
        for line in OPENCLAW_CRON_RUNS.read_text(encoding="utf-8").splitlines():
            if line.strip():
                last = json.loads(line)
    except (json.JSONDecodeError, OSError) as exc:
        return f"Unable to read cron run history: {exc}"
    if not last:
        return f"Cron run history is empty: {OPENCLAW_CRON_RUNS}"

    status = last.get("status")
    resolved = (
        last.get("delivery", {}).get("resolved", {})
        if isinstance(last.get("delivery"), dict)
        else {}
    )
    thread = resolved.get("threadId")
    if thread:
        return f"Latest historical cron run status={status}; resolved Telegram threadId={thread}"
    return f"Latest historical cron run status={status}; no resolved Telegram threadId recorded"


def _openclaw_delivery_health() -> str:
    data = _load_json(OPENCLAW_CRON_JOBS)
    jobs = data.get("jobs") if isinstance(data.get("jobs"), list) else []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        if job.get("id") != OPENCLAW_CRON_JOB_ID and job.get("name") != OPENCLAW_CRON_JOB_NAME:
            continue
        delivery = job.get("delivery") if isinstance(job.get("delivery"), dict) else {}
        failure = job.get("failureAlert") if isinstance(job.get("failureAlert"), dict) else {}
        delivery_thread = delivery.get("threadId") or delivery.get("messageThreadId")
        failure_thread = failure.get("threadId") or failure.get("messageThreadId")
        base = (
            "OpenClaw cron delivery: "
            f"{delivery.get('channel')}:{delivery.get('to')}"
            f"{':' + str(delivery_thread) if delivery_thread else ''}; "
            f"failureAlert={failure.get('channel')}:{failure.get('to')}"
            f"{':' + str(failure_thread) if failure_thread else ''}"
        )
        if delivery_thread and failure_thread and str(delivery_thread) != str(failure_thread):
            return f"{base}; warning: delivery/failureAlert thread mismatch"
        if not delivery_thread:
            return f"{base}; warning: no explicit threadId, fallback may drift to a stale Telegram topic. {_latest_run_delivery_line()}"
        return f"{base}. {_latest_run_delivery_line()}"
    return f"{OPENCLAW_CRON_JOB_NAME} cron not found in {OPENCLAW_CRON_JOBS}"


def _runtime_health(tuning: dict) -> str:
    checks = [
        f"Workdir exists: {WORKDIR.exists()} ({WORKDIR})",
        f"Tracker exists: {TRACKER.exists()} ({TRACKER})",
        f"Keyword pack exists: {KEYWORD_PACK.exists()} ({KEYWORD_PACK})",
        f"Tuning config: {tuning.get('config_status')}",
        f"Feedback log exists: {FEEDBACK_LOG.exists()} ({FEEDBACK_LOG})",
        _openclaw_delivery_health(),
    ]
    return "\n".join(f"- {line}" for line in checks)


def run_plan(now: dt.datetime, tuning: dict | None = None) -> str:
    tuning = tuning or _load_tuning()
    day_offset = now.toordinal()
    budgets = tuning.get("budgets") if isinstance(tuning.get("budgets"), dict) else {}
    company_limit = int(budgets.get("target_companies_per_run") or 6)

    companies = _target_company_rows()
    company_focus = _cycle(companies, company_limit, day_offset)

    validated_min = budgets.get("validated_target_min", 3)
    validated_max = budgets.get("validated_target_max", 8)
    manual_min = budgets.get("manual_check_target_min", 15)
    manual_max = budgets.get("manual_check_target_max", 25)
    freshness = tuning.get("freshness") if isinstance(tuning.get("freshness"), dict) else {}
    depth = tuning.get("depth") if isinstance(tuning.get("depth"), dict) else {}
    scoring = tuning.get("scoring") if isinstance(tuning.get("scoring"), dict) else {}

    lines = [
        "## Daily run plan",
        (
            f"Strategy preset: {tuning.get('preset')} | "
            "config-driven coverage; Northern Germany first unless preset says otherwise."
        ),
        f"Validated target: save {validated_min}-{validated_max} score>={scoring.get('save_threshold', 0.5)} recommendations when enough exist.",
        f"Manual-check target: report {manual_min}-{manual_max} plausible blocked/snippet-only links when enough exist.",
        (
            "Freshness target: "
            f"prefer postings <= {freshness.get('prefer_posted_within_days', 14)} days; "
            f"recheck seen leads after {freshness.get('recheck_seen_after_days', 21)} days; "
            f"allow unknown date={freshness.get('allow_unknown_date', True)}."
        ),
        (
            "Validation depth: "
            f"{depth.get('validation_mode', 'standard')}; "
            f"max full-JD validations={depth.get('max_full_jd_validations', 10)}; "
            f"require full JD for validated save={depth.get('require_full_jd_for_save', True)}."
        ),
        "Lead decision policy: run lead_decision.py before validated saves; source/freshness policy gates the score.",
        "",
        _daily_search_plan(now, tuning),
    ]

    lines.extend(["", "### Target-company focus"])
    if company_focus:
        for row in company_focus:
            lines.append(
                f"- {row['company']} — {row['region']} — hints: {row['hints']}"
            )
    else:
        lines.append("- Target company file unavailable; use direct company/ATS pages when discovered.")

    lines.extend(
        [
            "",
            "### Hard filters and caveats",
            "- Exclude Postdoc programs from validated recommendations.",
            "- PhD/Promotion-required roles should be manual-check or rejected, not validated saves.",
            "- Wissenschaftliche*r Mitarbeiter*in roles are allowed only when no Postdoc/PhD requirement and fit/location are plausible.",
            "- C++/C#-first titles or required primary stack stay excluded or manual-check; Python/CV-primary with secondary C++ is allowed with a risk note.",
            "- Access-limited, snippet-only, source-failure, or aggregate-only candidates stay Manual-check unless a full JD validates source confidence, seniority, location, and risk blockers.",
            "",
            "### Recent explicit feedback",
            _feedback_summary(tuning),
            "",
            "### Recent diagnostics to apply",
            _recent_diagnostics(),
            "",
            "### Recent recommendation keys to avoid duplicating",
            _recent_recommendation_keys(),
        ]
    )
    return "\n".join(lines)


def _tracker_fingerprint() -> dict:
    if not TRACKER.exists():
        return {
            "tracker_exists": False,
            "tracker_mtime": None,
            "application_count": None,
            "recommendation_count": None,
        }

    code = f"""
from openpyxl import load_workbook
from pathlib import Path
p = Path({str(TRACKER)!r})
wb = load_workbook(p, data_only=True)
apps = wb['Applications']
recs = wb['Recommendations']
app_count = sum(1 for r in range(2, apps.max_row + 1) if apps.cell(r, 1).value)
rec_count = sum(1 for r in range(2, recs.max_row + 1) if recs.cell(r, 1).value)
print(app_count, rec_count)
"""
    ok, out = run(["uv", "run", "python", "-c", code], timeout=60)
    app_count = rec_count = None
    if ok:
        parts = out.split()
        if len(parts) == 2:
            app_count, rec_count = int(parts[0]), int(parts[1])
    return {
        "tracker_exists": True,
        "tracker_mtime": TRACKER.stat().st_mtime,
        "application_count": app_count,
        "recommendation_count": rec_count,
    }


def workbook_snapshot() -> str:
    if not TRACKER.exists():
        return f"Tracker missing: {TRACKER}"

    code = f"""
from collections import Counter
from openpyxl import load_workbook
from pathlib import Path
p = Path({str(TRACKER)!r})
wb = load_workbook(p, data_only=True)

apps = wb['Applications']
app_headers = [apps.cell(1, c).value for c in range(1, apps.max_column + 1)]
app_idx = {{h: i + 1 for i, h in enumerate(app_headers)}}
app_status = Counter()
for r in range(2, apps.max_row + 1):
    app_id = apps.cell(r, app_idx['app_id']).value
    if app_id:
        app_status[apps.cell(r, app_idx['status']).value or 'blank'] += 1

recs = wb['Recommendations']
headers = [recs.cell(1, c).value for c in range(1, recs.max_column + 1)]
idx = {{h: i + 1 for i, h in enumerate(headers)}}
rec_status = Counter()
source_type = Counter()
recent = []
for r in range(2, recs.max_row + 1):
    rec_id = recs.cell(r, idx['rec_id']).value
    if not rec_id:
        continue
    rec_status[recs.cell(r, idx['status']).value or 'blank'] += 1
    source_type[recs.cell(r, idx['source_type']).value or 'blank'] += 1
    recent.append({{
        'rec_id': rec_id,
        'company': recs.cell(r, idx['company']).value,
        'position': recs.cell(r, idx['position']).value,
        'location': recs.cell(r, idx['location']).value,
        'score': recs.cell(r, idx['score']).value,
        'status': recs.cell(r, idx['status']).value,
        'domain_fit': recs.cell(r, idx['domain_fit']).value,
        'risk_flags': recs.cell(r, idx['risk_flags']).value,
    }})

print(f"Applications: {{sum(app_status.values())}} | by status: {{dict(app_status)}}")
print(f"Recommendations: {{sum(rec_status.values())}} | by status: {{dict(rec_status)}}")
print(f"Recommendation source types: {{dict(source_type.most_common(8))}}")
print()
print('Recent recommendations, newest last:')
for row in recent[-{RECENT_REC_LIMIT}:]:
    print(f"{{row['rec_id']}} | {{row['status']}} | {{row['company']}} | {{row['position']}} | {{row['location']}} | score={{row['score']}} | domain={{row['domain_fit']}} | risk={{row['risk_flags']}}")
"""
    ok, out = run(["uv", "run", "python", "-c", code])
    return out if ok else out


def _should_skip_duplicate(fingerprint: dict, now: dt.datetime) -> str | None:
    force_values = {
        os.environ.get("OPENCLAW_CRON_FORCE_WAKE", ""),
        os.environ.get("HERMES_CRON_FORCE_WAKE", ""),
    }
    if any(value.strip().lower() in {"1", "true", "yes"} for value in force_values):
        return None

    state = _load_state()
    last_at_raw = state.get("last_wake_at")
    last_fp = state.get("last_fingerprint")
    if not last_at_raw or not isinstance(last_fp, dict):
        return None

    try:
        last_at = dt.datetime.fromisoformat(last_at_raw)
        if last_at.tzinfo is None:
            last_at = last_at.replace(tzinfo=ZoneInfo("Europe/Berlin"))
    except ValueError:
        return None

    age_minutes = (now - last_at).total_seconds() / 60
    if age_minutes > DUPLICATE_WINDOW_MINUTES:
        return None

    comparable_keys = (
        "tracker_exists",
        "tracker_mtime",
        "application_count",
        "recommendation_count",
    )
    if all(fingerprint.get(k) == last_fp.get(k) for k in comparable_keys):
        return (
            f"duplicate tick within {DUPLICATE_WINDOW_MINUTES}m; tracker unchanged since "
            f"{last_at.isoformat(timespec='minutes')}"
        )
    return None


def _emit(lines: list[str], wake: bool, reason: str = "") -> None:
    print("\n".join(lines))
    gate = {"wakeAgent": wake}
    if reason:
        gate["reason"] = reason
    print(json.dumps(gate, ensure_ascii=False))


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Collect pre-run context for job scouting cron")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print context without updating the duplicate wake-gate state.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ignore duplicate wake-gate state for this invocation.",
    )
    args = parser.parse_args(argv)

    now = dt.datetime.now(ZoneInfo("Europe/Berlin"))
    tuning = _load_tuning()
    fingerprint = _tracker_fingerprint()

    if not fingerprint["tracker_exists"]:
        _emit(
            [
                "# Daily job-scouting preflight",
                f"Run timestamp: {now.isoformat(timespec='seconds')}",
                f"Workspace: {WORKDIR}",
                "",
                "## Runtime health",
                _runtime_health(tuning),
                f"Tracker missing: {TRACKER}",
            ],
            wake=True,
            reason="tracker_missing",
        )
        return

    skip_reason = None if args.force or args.dry_run else _should_skip_duplicate(fingerprint, now)
    if skip_reason:
        _emit(
            [
                "# Daily job-scouting preflight",
                f"Run timestamp: {now.isoformat(timespec='seconds')}",
                "",
                "## Runtime health",
                _runtime_health(tuning),
                "Agent run skipped by wakeAgent gate.",
            ],
            wake=False,
            reason=skip_reason,
        )
        return

    validate_ok, validate_out = run(["uv", "run", "python", "scripts/tracker_ops.py", "validate"])
    dedupe_ok, dedupe_out = run(
        ["uv", "run", "python", "scripts/tracker_ops.py", "dedupe-recs", "--dry-run"]
    )
    kw_ok, kw_out = run(["uv", "run", "python", "scripts/search_keywords.py"], timeout=30)
    if not kw_ok:
        kw_out = kw_out or (
            "Keyword pack not loaded. Copy your private keyword pack to "
            "docs/agent_job_search_keywords.json in the workspace."
        )

    lines = [
        "# Daily job-scouting preflight context",
        f"Run timestamp: {now.isoformat(timespec='seconds')}",
        f"Workspace: {WORKDIR}",
        f"Tracker: {TRACKER}",
        "",
        "## Runtime health",
        _runtime_health(tuning),
        "",
        format_tuning_summary(tuning) if format_tuning_summary else f"## Tuning profile\nUnavailable: {TUNING_IMPORT_ERROR}",
        "",
        "## Tracker snapshot",
        workbook_snapshot(),
        "",
        "## Tracker validation",
        validate_out,
        "",
        "## Recommendation duplicate dry-run",
        dedupe_out,
        "",
        run_plan(now, tuning),
        "",
        kw_out,
        "",
        "## Operational reminder",
        textwrap.dedent(
            """
            Use this snapshot to steer search away from already-tracked companies/titles.
            Before saving new validated recommendations, still deduplicate against the live tracker;
            after saving, run tracker validation and duplicate dry-run again.
            """
        ).strip(),
    ]

    wake_reason = "ok"
    if not validate_ok:
        wake_reason = "tracker_validate_failed"

    _emit(lines, wake=True, reason=wake_reason)
    if not args.dry_run:
        _save_state(
            {
                "last_wake_at": now.isoformat(timespec="seconds"),
                "last_fingerprint": fingerprint,
            }
        )


if __name__ == "__main__":
    main()
